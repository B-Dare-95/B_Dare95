import os
import threading
import time
import subprocess
import tkinter as tk
from tkinter import filedialog, messagebox
from PIL import ImageGrab
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage

# ==========================================
# SETTINGS
# ==========================================

SNIPASTE_PATH = r"C:\Users\moham\AppData\Local\Microsoft\WindowsApps\Snipaste.exe"

# ==========================================
# SELECT / CREATE EXCEL FILE
# ==========================================

root = tk.Tk()
root.withdraw()

save_folder = filedialog.askdirectory(
    title="Select where to save Issue Logger.xlsx"
)

if not save_folder:
    raise SystemExit("No folder selected.")

excel_path = os.path.join(save_folder, "Issue Logger.xlsx")

# Create workbook if it doesn't exist
if not os.path.exists(excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Issues"

    ws["A1"] = "Timestamp"
    ws["B1"] = "Comment"
    ws["C1"] = "Screenshot"

    wb.save(excel_path)

# ==========================================
# MAIN UI
# ==========================================

app = tk.Tk()
app.title("Issue Logger")
app.geometry("500x300")

screenshot_ready = False
latest_image_path = None

# ------------------------------------------
# COMMENT LABEL
# ------------------------------------------

label = tk.Label(app, text="Write your comment:")
label.pack(pady=(15, 5))

# ------------------------------------------
# COMMENT TEXTBOX
# ------------------------------------------

comment_box = tk.Text(app, height=8, width=55)
comment_box.pack(pady=5)

# ------------------------------------------
# TAKE SCREENSHOT
# ------------------------------------------

def take_screenshot():
    # Hide immediately
    app.withdraw()
    app.update_idletasks()
    app.update()

    # Start Snipaste
    process = subprocess.Popen([
        SNIPASTE_PATH,
        "snip"
    ])

    # Wait for Snipaste to finish, then restore UI
    def watch_snipaste():
        process.wait()
        app.after(0, app.deiconify)

    threading.Thread(target=watch_snipaste, daemon=True).start()

# ------------------------------------------
# SAVE ISSUE
# ------------------------------------------

def save_issue():
    global latest_image_path

    comment = comment_box.get("1.0", tk.END).strip()

    if not comment:
        messagebox.showerror("Error", "Please enter a comment.")
        return

    # Get image from clipboard
    img = ImageGrab.grabclipboard()

    if img is None:
        messagebox.showerror(
            "Error",
            "No image found in clipboard.\n\n"
            "Please take a screenshot with Snipaste first."
        )
        return

    # Load workbook
    wb = load_workbook(excel_path)
    ws = wb.active

    # Next row
    next_row = ws.max_row + 1

    # Timestamp
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")

    # Save screenshot temporarily
    images_folder = os.path.join(save_folder, "Issue_Images")
    os.makedirs(images_folder, exist_ok=True)

    image_filename = f"issue_{next_row}.png"
    image_path = os.path.join(images_folder, image_filename)

    img.save(image_path)

    # Write text data
    ws.cell(next_row, 1).value = timestamp
    ws.cell(next_row, 2).value = comment

    # Insert image into Excel
    excel_img = XLImage(image_path)
    excel_img.width = 300
    excel_img.height = 170

    ws.add_image(excel_img, f"C{next_row}")

    # Adjust row height
    ws.row_dimensions[next_row].height = 140

    # Adjust columns
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 45

    # Save workbook
    wb.save(excel_path)

    # Clear textbox
    comment_box.delete("1.0", tk.END)

    messagebox.showinfo(
        "Saved",
        "Issue saved successfully."
    )

# ------------------------------------------
# BUTTONS
# ------------------------------------------

btn_frame = tk.Frame(app)
btn_frame.pack(pady=20)

snip_btn = tk.Button(
    btn_frame,
    text="Take Screenshot",
    width=20,
    height=2,
    command=take_screenshot
)

snip_btn.grid(row=0, column=0, padx=10)

save_btn = tk.Button(
    btn_frame,
    text="Save Issue",
    width=20,
    height=2,
    command=save_issue
)

save_btn.grid(row=0, column=1, padx=10)

# ==========================================
# RUN APP
# ==========================================

app.mainloop()